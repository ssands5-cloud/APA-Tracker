"""End-to-end: GraphQL response -> ingestion -> SQLite -> Excel export.

Runs the whole chain the live sync runs, with the network call replaced by
an invented response (see tests/test_graphql_scraper.py on fixtures). This
is the test that would have caught `database.ingest` being un-importable:
it exercises the real ingestion functions against a real database.
"""

from __future__ import annotations

import pytest
from sqlalchemy import create_engine
from sqlalchemy.orm import Session

from database.models import Base, Match, Player, Team
from scheduler.graphql_sync import ingest_team_data
from tests.test_graphql_scraper import TEAM_DATA


@pytest.fixture
def db():
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    with Session(engine) as session:
        yield session


class TestIngestTeamData:
    def test_team_roster_and_schedule_all_land_in_the_database(self, db):
        counts = ingest_team_data(db, TEAM_DATA)

        assert counts == {
            "roster": 1, "standings": 1, "matches_seen": 1, "matches_new": 1,
            "matches_updated": 0, "byes": 0, "unscored": 0,
        }
        team = db.query(Team).one()
        assert team.external_id == "13082948"
        assert team.name == "Chalk It Up"

        player = db.query(Player).one()
        assert player.name == "Alex R."
        # These columns did not exist before; the values were silently dropped.
        assert (player.matches_won, player.matches_played) == (8, 10)
        assert player.win_pct == 0.8
        assert player.team_id == team.id

        match = db.query(Match).one()
        assert match.external_id == "555001"
        assert match.away_team_name == "Rack Attack"
        assert match.match_date == "2026-06-04T19:00:00Z"

    def test_rerunning_does_not_duplicate_matches(self, db):
        ingest_team_data(db, TEAM_DATA)
        second = ingest_team_data(db, TEAM_DATA)

        assert second["matches_new"] == 0, "a re-sync must not re-insert known matches"
        assert second["matches_updated"] == 1
        assert db.query(Match).count() == 1
        assert db.query(Player).count() == 1

    def test_a_score_arriving_later_updates_the_existing_match(self, db):
        """The schedule is published before the season, so almost every match
        is first seen unplayed. Skipping known matches meant a result could
        never arrive."""
        unplayed = {
            "team": TEAM_DATA["team"], "roster": [],
            "schedule": [{"id": 700, "week": 2, "status": "SCHEDULED", "isScored": False,
                          "home": {"id": 13082948, "name": "Chalk It Up"},
                          "away": {"id": 2, "name": "Rack Attack"}}],
        }
        ingest_team_data(db, unplayed)
        match = db.query(Match).one()
        assert (match.home_score, match.away_score) == (None, None)
        assert match.is_scored is False

        played = {
            "team": TEAM_DATA["team"], "roster": [],
            "schedule": [dict(unplayed["schedule"][0], status="COMPLETED", isScored=True,
                              isFinalized=True,
                              results=[{"homeAway": "HOME", "points": {"total": 3}},
                                       {"homeAway": "AWAY", "points": {"total": 2}}])],
        }
        counts = ingest_team_data(db, played)

        assert counts["matches_new"] == 0 and counts["matches_updated"] == 1
        db.expire_all()
        match = db.query(Match).one()
        assert (match.home_score, match.away_score) == (3, 2)
        assert match.is_scored is True and match.is_finalized is True
        assert match.week == 2

    def test_partially_scored_match_keeps_the_missing_side_null(self, db):
        data = {
            "team": TEAM_DATA["team"], "roster": [],
            "schedule": [{"id": 701, "isScored": True, "isFinalized": False,
                          "home": {"id": 13082948, "name": "Chalk It Up"},
                          "away": {"id": 2, "name": "Rack Attack"},
                          "results": [{"homeAway": "HOME", "points": {"total": 2}}]}],
        }
        ingest_team_data(db, data)
        match = db.query(Match).one()
        assert match.home_score == 2
        assert match.away_score is None, "a missing side is unknown, not zero"
        assert match.is_finalized is False

    def test_bye_week_is_recorded_as_a_bye(self, db):
        data = {
            "team": TEAM_DATA["team"],
            "roster": [],
            "schedule": [{"id": 900, "week": 5, "isBye": True,
                          "home": {"id": 13082948, "name": "Chalk It Up"}, "away": None}],
        }
        counts = ingest_team_data(db, data)

        assert counts["byes"] == 1
        assert db.query(Match).one().away_team_name == "BYE"

    def test_schedule_entry_with_no_id_is_skipped_not_crashed(self, db):
        data = {"team": TEAM_DATA["team"], "roster": [], "schedule": [{"week": 3}]}
        counts = ingest_team_data(db, data)

        assert counts["matches_new"] == 0
        assert db.query(Match).count() == 0

    def test_unscored_matches_are_counted_separately(self, db):
        data = {
            "team": TEAM_DATA["team"],
            "roster": [],
            "schedule": [{"id": 901, "week": 6, "isScored": False, "status": "SCHEDULED",
                          "home": {"id": 1, "name": "A"}, "away": {"id": 2, "name": "B"}}],
        }
        assert ingest_team_data(db, data)["unscored"] == 1


class TestExcelExportSeesLiveData:
    def test_export_contains_the_ingested_players(self, db, tmp_path):
        pytest.importorskip("pandas")
        pytest.importorskip("openpyxl")
        import openpyxl

        from ui.export_excel import export_to_excel

        ingest_team_data(db, TEAM_DATA)
        out = tmp_path / "apa_stats.xlsx"
        path = export_to_excel(db, {"export": {"excel_output_path": str(out)}})

        workbook = openpyxl.load_workbook(path)
        assert workbook.sheetnames == ["Standings", "Player Stats"]

        rows = [[cell.value for cell in row] for row in workbook["Player Stats"].iter_rows()]
        header, alex = rows[0], rows[1]
        record = dict(zip(header, alex))

        assert record["Player"] == "Alex R."
        # The regression: roster-sourced players used to export as 0-0.
        assert record["Matches"] == 10
        assert record["Wins"] == 8
        assert record["Losses"] == 2
        assert record["Win %"] == 0.8
        assert record["Skill Level"] == 5
        assert record["Source"] == "roster totals"

    def test_standings_sheet_shows_our_live_rank(self, db, tmp_path):
        pytest.importorskip("pandas")
        pytest.importorskip("openpyxl")
        import openpyxl

        from ui.export_excel import export_to_excel

        ingest_team_data(db, TEAM_DATA)
        path = export_to_excel(db, {"export": {"excel_output_path": str(tmp_path / "s.xlsx")}})

        rows = [[c.value for c in r] for r in openpyxl.load_workbook(path)["Standings"].iter_rows()]
        record = dict(zip(rows[0], rows[1]))
        assert record["Team"] == "Chalk It Up"
        assert record["Rank"] == 3

    def test_player_with_no_data_at_all_is_not_reported_as_a_loss(self, db, tmp_path):
        """0 played must not render as 0 wins / 0 losses with no explanation."""
        pytest.importorskip("pandas")
        from ui.export_excel import _player_stats_dataframe
        from database.ingest import upsert_player

        upsert_player(db, "P9", "Newcomer")
        frame = _player_stats_dataframe(db)
        assert frame.iloc[0]["Source"] == "no data"


class TestRealDivisionStandings:
    """With the division query captured (2026-09-03), standings are the real
    table for every team, not our own row derived from our own schedule."""

    DIVISION = {
        "id": 436670,
        "teams": [
            {"id": 1, "name": "Rack Attack", "standing": 1, "sessionTotalPoints": 72},
            {"id": 13082948, "name": "Chalk It Up", "standing": 3, "sessionTotalPoints": 57},
            {"id": 9, "name": "Corner Pocket", "standing": 5, "sessionTotalPoints": 41},
        ],
    }

    def test_all_division_teams_are_snapshotted(self, db):
        from database.models import StandingsSnapshot

        data = dict(TEAM_DATA, division=self.DIVISION)
        counts = ingest_team_data(db, data)

        assert counts["standings"] == 3, "the whole division, not just our team"
        rows = db.query(StandingsSnapshot).all()
        assert {r.team_name for r in rows} == {"Rack Attack", "Chalk It Up", "Corner Pocket"}
        ours = next(r for r in rows if r.team_name == "Chalk It Up")
        assert (ours.rank, ours.points) == (3, 57)

    def test_falls_back_to_our_team_when_no_division_data(self, db):
        """A config without a division id still gets its own snapshot."""
        counts = ingest_team_data(db, TEAM_DATA)
        assert counts["standings"] == 1

    def test_the_excel_standings_sheet_shows_the_division(self, db, tmp_path):
        pytest.importorskip("openpyxl")
        import openpyxl

        from ui.export_excel import export_to_excel

        ingest_team_data(db, dict(TEAM_DATA, division=self.DIVISION))
        path = export_to_excel(db, {"export": {"excel_output_path": str(tmp_path / "d.xlsx")}})

        sheet = openpyxl.load_workbook(path)["Standings"]
        rows = [[c.value for c in r] for r in sheet.iter_rows()]
        header, data_rows = rows[0], rows[1:]
        assert len(data_rows) == 3
        # Ordered by rank, as latest_standings() sorts them.
        assert [dict(zip(header, r))["Team"] for r in data_rows] == [
            "Rack Attack", "Chalk It Up", "Corner Pocket",
        ]
