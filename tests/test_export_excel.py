"""Tests for ui.export_excel -- first coverage this module has had.

Mirrors tests/test_export_json.py's approach: seed a small in-memory
database through the real ingest functions, then check the actual
workbook produced, not just that export_to_excel() didn't raise.
"""

from __future__ import annotations

import openpyxl
import pytest
from sqlalchemy import create_engine
from sqlalchemy.orm import Session

from database.ingest import (
    ingest_eight_ball_stats,
    ingest_match,
    ingest_match_scores,
    ingest_player_team_history,
    ingest_standings,
    upsert_player,
    upsert_team,
)
from database.models import Base
from ui.export_excel import export_to_excel

EXPECTED_SHEETS = {"Standings", "Player Stats", "Career Stats", "Team History", "Skill Level History"}


@pytest.fixture
def db():
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    with Session(engine) as session:
        yield session


def _export(db, tmp_path):
    config = {"export": {"excel_output_path": str(tmp_path / "out.xlsx")}}
    path = export_to_excel(db, config)
    return openpyxl.load_workbook(path)


class TestEmptyDatabase:
    """A fresh checkout runs the demo/sync before anything is ingested --
    every sheet must exist, even with only a header row."""

    def test_all_four_sheets_exist(self, db, tmp_path):
        wb = _export(db, tmp_path)
        assert set(wb.sheetnames) == EXPECTED_SHEETS

    def test_career_stats_and_team_history_are_empty_not_missing(self, db, tmp_path):
        wb = _export(db, tmp_path)
        assert wb["Career Stats"].max_row == 1  # header only
        assert wb["Team History"].max_row == 1
        assert wb["Skill Level History"].max_row == 1


class TestSeededData:
    @pytest.fixture
    def seeded_db(self, db):
        team = upsert_team(db, "T1", "Mark It Up")
        player = upsert_player(db, "3349374", "Paul Smith", team)
        ingest_eight_ball_stats(db, player, {
            "eight_ball_matches_won": 64, "eight_ball_matches_played": 129,
            "eight_ball_cla": 1, "eight_ball_defensive_shot_avg": 1.26,
            "eight_ball_match_count_for_last_two_yrs": 123, "eight_ball_last_played": "2026-08-31",
        })
        ingest_player_team_history(db, player, [{
            "is_current": True, "team_name": "Mark It Up", "division_id": "436670",
            "is_tournament": False, "session_name": "2026 Summer", "nick_name": "Paulie",
            "skill_level": 4, "rank": None, "matches_won": 2, "matches_played": 2,
        }])
        ingest_match(db, match_id="M1", home_team_id="T1", away_team_id="T2",
                     home_team_name="Mark It Up", away_team_name="Rack Attack",
                     status="COMPLETED", home_score=18, away_score=12)
        ingest_match_scores(db, "M1", [
            {"player_id": "3349374", "player_name": "Paul Smith", "team_id": "T1",
             "result": "W", "points_earned": 6},
        ])
        ingest_standings(db, [{"team_name": "Mark It Up", "rank": 1, "points": 45}])
        return db

    def test_career_stats_sheet_has_the_real_columns_and_values(self, seeded_db, tmp_path):
        wb = _export(seeded_db, tmp_path)
        ws = wb["Career Stats"]
        headers = [c.value for c in ws[1]]
        assert headers == [
            "Player", "Format", "Matches Won", "Matches Played", "CLA",
            "Defensive Shot Avg", "Matches (Last 2 Yrs)", "Last Played",
        ]
        row = [c.value for c in ws[2]]
        assert row == ["Paul Smith", "EIGHT", 64, 129, 1, 1.26, 123, "2026-08-31"]

    def test_team_history_sheet_has_the_real_columns_and_values(self, seeded_db, tmp_path):
        wb = _export(seeded_db, tmp_path)
        ws = wb["Team History"]
        headers = [c.value for c in ws[1]]
        assert headers == [
            "Player", "Current", "Team", "Division", "Tournament", "Session",
            "Nickname", "Skill Level", "Rank", "Matches Won", "Matches Played",
        ]
        row = [c.value for c in ws[2]]
        assert row == [
            "Paul Smith", True, "Mark It Up", "436670", False, "2026 Summer",
            "Paulie", 4, None, 2, 2,
        ]

    def test_standings_and_player_stats_still_work_alongside_the_new_sheets(self, seeded_db, tmp_path):
        wb = _export(seeded_db, tmp_path)
        assert wb["Standings"].max_row == 2  # header + 1 team
        assert wb["Player Stats"].max_row == 2  # header + 1 player

    def test_player_stats_shows_which_team_the_row_is_for(self, seeded_db, tmp_path):
        """A player on two teams during a season legitimately gets two rows
        here -- without a Team column that looked like an unexplained
        duplicate. Real report from Paul: seeing his own name twice, which
        turned out to be exactly this, not a bug."""
        wb = _export(seeded_db, tmp_path)
        rows = [[c.value for c in r] for r in wb["Player Stats"].iter_rows()]
        header, paul = rows[0], rows[1]
        record = dict(zip(header, paul))
        assert record["Team"] == "Mark It Up"

    def test_skill_level_history_sheet_has_the_real_columns_and_values(self, db, tmp_path):
        """Not a new extraction step -- ingest_match_scores already writes
        PlayerMatch.skill_level per match (seeded_db above doesn't set it,
        so this seeds its own match with one)."""
        upsert_team(db, "T1", "Mark It Up")
        upsert_team(db, "T2", "Rack Attack")
        ingest_match(db, match_id="M1", home_team_id="T1", away_team_id="T2",
                     home_team_name="Mark It Up", away_team_name="Rack Attack",
                     week=1, match_date="2026-06-01", status="COMPLETED",
                     home_score=6, away_score=3)
        ingest_match_scores(db, "M1", [
            {"player_id": "3349374", "player_name": "Paul Smith", "team_id": "T1",
             "skill_level": 5, "result": "W", "points_earned": 6},
        ])
        wb = _export(db, tmp_path)
        ws = wb["Skill Level History"]
        headers = [c.value for c in ws[1]]
        assert headers == ["Player Name", "Player ID", "Week", "Skill Level", "Match Date", "Source"]
        row = [c.value for c in ws[2]]
        assert row == ["Paul Smith", "3349374", 1, 5, "2026-06-01", "scoresheet"]

    def test_player_never_rostered_shows_a_blank_team_not_a_crash(self, db, tmp_path):
        """ingest_match_scores() never assigns a team (only upsert_roster()
        does) -- a player known only from a scoresheet has no team to show."""
        upsert_team(db, "T1", "Chalk It Up")
        upsert_team(db, "T2", "Rack Attack")
        ingest_match(db, match_id="M1", home_team_id="T1", away_team_id="T2",
                     home_team_name="Chalk It Up", away_team_name="Rack Attack",
                     status="COMPLETED", home_score=6, away_score=3)
        ingest_match_scores(db, "M1", [
            {"player_id": "P9", "player_name": "Opponent Guy", "team_id": "T2",
             "result": "L", "points_earned": 3},
        ])
        wb = _export(db, tmp_path)
        rows = [[c.value for c in r] for r in wb["Player Stats"].iter_rows()]
        record = dict(zip(rows[0], rows[1]))
        # An empty string round-trips through openpyxl as a blank cell (None).
        assert not record["Team"]
