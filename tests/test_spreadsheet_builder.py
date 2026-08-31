"""
Integration tests for SpreadsheetBuilder using an in-memory SQLite database.
"""

from __future__ import annotations

import openpyxl
import pytest

from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from database.models import Base, Player, PlayerMatch, StandingsSnapshot, Team
from analytics.spreadsheet_builder import SpreadsheetBuilder, build_from_db


@pytest.fixture()
def db_session(tmp_path):
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    Session = sessionmaker(bind=engine)
    session = Session()

    # Teams
    t1 = Team(external_id="team-1", name="Alpha Squad")
    t2 = Team(external_id="team-2", name="Beta Force")
    session.add_all([t1, t2])
    session.flush()

    # Players
    p1 = Player(external_id="p-100", name="Alice Smith", skill_level=5, team_id=t1.id)
    p2 = Player(external_id="p-200", name="Bob Jones", skill_level=3, team_id=t1.id)
    p3 = Player(external_id="p-300", name="Carol White", skill_level=4, team_id=t2.id)
    session.add_all([p1, p2, p3])
    session.flush()

    # PlayerMatches
    session.add_all([
        PlayerMatch(player_id=p1.id, match_date="2026-01-10", opponent="Carol White",
                    skill_level=5, points_earned=3.0, result="W"),
        PlayerMatch(player_id=p1.id, match_date="2026-01-17", opponent="Bob Jones",
                    skill_level=5, points_earned=2.0, result="L"),
        PlayerMatch(player_id=p1.id, match_date="2026-01-24", opponent="Carol White",
                    skill_level=5, points_earned=3.5, result="W"),
        PlayerMatch(player_id=p2.id, match_date="2026-01-17", opponent="Alice Smith",
                    skill_level=3, points_earned=4.0, result="W"),
        PlayerMatch(player_id=p3.id, match_date="2026-01-10", opponent="Alice Smith",
                    skill_level=4, points_earned=1.0, result="L"),
    ])

    # StandingsSnapshot
    from datetime import datetime
    ts = datetime(2026, 1, 31)
    session.add_all([
        StandingsSnapshot(captured_at=ts, team_name="Alpha Squad", rank=1, wins=5, losses=2, points=18.0),
        StandingsSnapshot(captured_at=ts, team_name="Beta Force", rank=2, wins=3, losses=4, points=11.0),
    ])
    session.commit()
    yield session
    session.close()


def test_build_all_sheets_no_errors(db_session, tmp_path):
    """SpreadsheetBuilder should complete without raising."""
    builder = SpreadsheetBuilder(db_session)
    builder.build_all_sheets()
    out = builder.export_to_excel(tmp_path / "test_output.xlsx")
    assert out.exists()


def test_workbook_has_expected_sheets(db_session, tmp_path):
    builder = SpreadsheetBuilder(db_session)
    builder.build_all_sheets()
    out = builder.export_to_excel(tmp_path / "test2.xlsx")
    wb = openpyxl.load_workbook(str(out))
    assert "PLAYER_LIFETIME" in wb.sheetnames
    assert "TEAM_STATS" in wb.sheetnames
    assert "HEAD_TO_HEAD" in wb.sheetnames
    assert "SEASON_SUMMARY" in wb.sheetnames


def test_player_lifetime_row_count(db_session, tmp_path):
    builder = SpreadsheetBuilder(db_session)
    builder.build_all_sheets()
    out = builder.export_to_excel(tmp_path / "test3.xlsx")
    wb = openpyxl.load_workbook(str(out))
    ws = wb["PLAYER_LIFETIME"]
    data_rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
    assert len(data_rows) == 3  # 3 players


def test_build_from_db_convenience(db_session, tmp_path):
    out = build_from_db(db_session, tmp_path / "convenience.xlsx")
    assert out.exists()


def test_team_stats_win_pct(db_session, tmp_path):
    from analytics.sheet_generators import TeamStatsSheet
    gen = TeamStatsSheet(db_session)
    rows = gen.rows()
    alpha = next(r for r in rows if r["Team Name"] == "Alpha Squad")
    assert alpha["Matches Won"] == 5
    assert alpha["Win %"] == pytest.approx(5 / 7, rel=1e-3)
