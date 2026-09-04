"""Smoke test for scripts/build_demo.py.

Not a unit test of ingestion logic -- that's covered elsewhere (see
test_viewer_sync_fixture.py, test_division_standings_fixture.py,
test_match_detail_fixture.py). This just proves the demo script still wires
those pieces together correctly end to end and produces a real workbook,
so a signature change in any of them fails loudly here instead of only
being noticed the next time someone runs the demo by hand.
"""

from __future__ import annotations

import json
import subprocess
import sys
from pathlib import Path

import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parent.parent


def test_build_demo_produces_a_readable_workbook(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    result = subprocess.run(
        [sys.executable, str(PROJECT_ROOT / "scripts" / "build_demo.py")],
        capture_output=True, text=True, cwd=tmp_path,
    )
    assert result.returncode == 0, result.stderr

    workbook_path = tmp_path / "exports" / "demo_apa_stats.xlsx"
    assert workbook_path.exists()

    wb = openpyxl.load_workbook(workbook_path)
    assert set(wb.sheetnames) == {
        "Standings", "Player Stats", "Career Stats", "Team History",
        "Skill Level History", "Matchups",
    }

    standings = wb["Standings"]
    assert standings.max_row > 1, "Standings sheet has no data rows"

    player_stats = wb["Player Stats"]
    assert player_stats.max_row > 1, "Player Stats sheet has no data rows"

    # P0-3: the Matchup Advantage Engine's real pipeline -- ingest_head_to_head
    # (from the match_detail fixture) -> analytics.matchup_builder.build_matchups()
    # -> export -- run end to end, not just a hand-crafted PlayerMatchup row
    # fed straight to ingest_matchups() the way tests/test_export_excel.py and
    # tests/test_export_json.py's matchup tests do. This is what would have
    # caught run_all_teams() never calling build_matchups() (P0-1): that gap
    # didn't break any existing test because none of them ran the real
    # ingest-to-export chain for matchups specifically.
    matchups_sheet = wb["Matchups"]
    assert matchups_sheet.max_row > 1, "Matchups sheet has no data rows"
    header = [c.value for c in matchups_sheet[1]]
    first_row = dict(zip(header, [c.value for c in matchups_sheet[2]]))
    assert first_row["Player"], "a real matchup row must name a real player"
    assert first_row["Opponent"], "a real matchup row must name a real opponent"
    assert first_row["Matchup Score"] is not None
    assert first_row["Confidence Score"] is not None

    json_path = tmp_path / "exports" / "demo_apa_data.json"
    assert json_path.exists()
    document = json.loads(json_path.read_text())
    assert document["matchups"], "the JSON export's matchups key must not be empty"
    json_row = document["matchups"][0]
    assert json_row["player"] and json_row["opponent"]
    assert json_row["matchup_score"] is not None
    assert json_row["confidence_score"] is not None
    # Same underlying player_matchups table, both exports -- the two rows
    # for a given pair must actually agree, not just both be "present".
    excel_pairs = {
        (row[0], row[1]) for row in matchups_sheet.iter_rows(min_row=2, values_only=True)
    }
    json_pairs = {(row["player"], row["opponent"]) for row in document["matchups"]}
    assert excel_pairs == json_pairs
