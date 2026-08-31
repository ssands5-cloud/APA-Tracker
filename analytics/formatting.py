"""
Excel formatting utilities for the APA analytics workbook.
"""

from __future__ import annotations

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# Colour palette
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
ALT_ROW_FILL = PatternFill("solid", fgColor="D6E4F0")
SUMMARY_FILL = PatternFill("solid", fgColor="2E75B6")

HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
SUMMARY_FONT = Font(color="FFFFFF", bold=True, size=12)
BOLD_FONT = Font(bold=True)


def apply_header_row(ws, headers: list[str]) -> None:
    """Write headers to row 1 with styling."""
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30


def auto_size_columns(ws, min_width: int = 10, max_width: int = 40) -> None:
    """Set column widths based on max content length."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 2, max_width))


def freeze_header(ws) -> None:
    """Freeze the first row so it stays visible while scrolling."""
    ws.freeze_panes = "A2"


def add_autofilter(ws, num_columns: int) -> None:
    """Add autofilter spanning all header columns.

    *num_columns* is the number of headers so the filter ref is correct
    even before data rows are appended.
    """
    last_col = get_column_letter(num_columns)
    ws.auto_filter.ref = f"A1:{last_col}1"


def style_sheet(ws, headers: list[str]) -> None:
    """Apply all standard formatting to a worksheet."""
    apply_header_row(ws, headers)
    freeze_header(ws)
    add_autofilter(ws, len(headers))
    auto_size_columns(ws)
